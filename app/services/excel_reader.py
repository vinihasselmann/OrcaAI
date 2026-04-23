"""Leitura e inspecao de planilhas Excel existentes (.xlsx)."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


class ExcelReaderError(Exception):
    """Erro de leitura/inspecao de planilha Excel."""


def _normalizar_texto(valor: Any) -> str:
    """Normaliza texto para comparacoes de cabecalho."""
    if valor is None:
        return ""
    return str(valor).strip().lower()


def _obter_aba(workbook: Workbook, nome_aba: str) -> Worksheet:
    """Retorna a aba pelo nome ou dispara erro amigavel."""
    if nome_aba not in workbook.sheetnames:
        raise ExcelReaderError(f"Aba nao encontrada: {nome_aba}")
    return workbook[nome_aba]


def abrir_workbook_existente(
    caminho: str | Path,
    *,
    read_only: bool = True,
    data_only: bool = False,
) -> Workbook:
    """Abre um arquivo .xlsx existente sem criar/recriar planilha."""
    caminho_path = Path(caminho)
    if not caminho_path.exists():
        raise ExcelReaderError(f"Arquivo Excel nao encontrado: {caminho_path}")
    if caminho_path.suffix.lower() != ".xlsx":
        raise ExcelReaderError(f"Formato invalido (esperado .xlsx): {caminho_path}")

    try:
        return load_workbook(
            filename=caminho_path,
            read_only=read_only,
            data_only=data_only,
        )
    except Exception as exc:  # pragma: no cover
        raise ExcelReaderError(f"Falha ao abrir workbook: {exc}") from exc


def listar_abas(workbook: Workbook) -> list[str]:
    """Lista os nomes das abas do workbook."""
    return list(workbook.sheetnames)


def ler_linha_inteira(
    workbook: Workbook,
    nome_aba: str,
    numero_linha: int,
) -> list[Any]:
    """Le todos os valores de uma linha da aba."""
    if numero_linha <= 0:
        raise ExcelReaderError("numero_linha deve ser maior que zero.")

    aba = _obter_aba(workbook, nome_aba)
    try:
        row_iter = aba.iter_rows(
            min_row=numero_linha,
            max_row=numero_linha,
            values_only=True,
        )
        return list(next(row_iter, ()))
    except Exception as exc:  # pragma: no cover
        raise ExcelReaderError(f"Falha ao ler linha {numero_linha}: {exc}") from exc


def localizar_cabecalhos(
    workbook: Workbook,
    nome_aba: str,
    cabecalhos_esperados: list[str] | None = None,
    *,
    linha_inicial: int = 1,
    linha_final: int = 30,
) -> dict[str, str | None]:
    """Localiza cabecalhos em uma aba.

    - Se `cabecalhos_esperados` for informado, retorna um mapeamento
      `{cabecalho_esperado: coordenada_ou_none}`.
    - Caso contrario, tenta inferir a linha de cabecalho mais provavel no range
      e retorna `{texto_cabecalho: coordenada}`.
    """
    if linha_inicial <= 0 or linha_final <= 0 or linha_final < linha_inicial:
        raise ExcelReaderError("Intervalo de linhas invalido para localizar cabecalhos.")

    aba = _obter_aba(workbook, nome_aba)

    if cabecalhos_esperados:
        resultado: dict[str, str | None] = {h: None for h in cabecalhos_esperados}
        esperados_norm = {_normalizar_texto(h): h for h in cabecalhos_esperados}

        for row in aba.iter_rows(min_row=linha_inicial, max_row=linha_final):
            for cell in row:
                texto = _normalizar_texto(cell.value)
                if texto in esperados_norm:
                    original = esperados_norm[texto]
                    resultado[original] = cell.coordinate
        return resultado

    melhor_linha: list[Any] = []
    melhor_indice = -1
    maior_qtd_texto = -1

    for idx, valores in enumerate(
        aba.iter_rows(min_row=linha_inicial, max_row=linha_final, values_only=True),
        start=linha_inicial,
    ):
        qtd_texto = sum(1 for v in valores if isinstance(v, str) and v.strip())
        if qtd_texto > maior_qtd_texto:
            maior_qtd_texto = qtd_texto
            melhor_linha = list(valores)
            melhor_indice = idx

    if melhor_indice < 0:
        return {}

    cabecalhos: dict[str, str] = {}
    for col_idx, valor in enumerate(melhor_linha, start=1):
        if not isinstance(valor, str) or not valor.strip():
            continue
        coordenada = f"{aba.cell(row=melhor_indice, column=col_idx).coordinate}"
        cabecalhos[str(valor).strip()] = coordenada
    return cabecalhos


def ler_valor_celula(
    workbook: Workbook,
    nome_aba: str,
    referencia_celula: str,
) -> Any:
    """Le o valor de uma celula especifica."""
    aba = _obter_aba(workbook, nome_aba)
    try:
        return aba[referencia_celula].value
    except Exception as exc:  # pragma: no cover
        raise ExcelReaderError(
            f"Referencia de celula invalida '{referencia_celula}': {exc}"
        ) from exc


def detectar_formula(
    workbook: Workbook,
    nome_aba: str,
    referencia_celula: str,
) -> bool:
    """Retorna True quando a celula contem formula Excel."""
    valor = ler_valor_celula(workbook, nome_aba, referencia_celula)
    return isinstance(valor, str) and valor.startswith("=")


def buscar_linhas_por_identificador(
    workbook: Workbook,
    nome_aba: str,
    identificador: str,
    *,
    colunas_identificadoras: list[str] | None = None,
    linha_cabecalho: int = 1,
    correspondencia_parcial: bool = True,
) -> list[dict[str, Any]]:
    """Busca linhas por identificador em colunas tecnicas.

    Exemplo de colunas padrao: `codigo_montagem`, `modelo`, `marca_tipo`.
    """
    if not identificador or not identificador.strip():
        raise ExcelReaderError("identificador nao pode ser vazio.")
    if linha_cabecalho <= 0:
        raise ExcelReaderError("linha_cabecalho deve ser maior que zero.")

    aba = _obter_aba(workbook, nome_aba)
    colunas_identificadoras = colunas_identificadoras or [
        "codigo_montagem",
        "modelo",
        "marca_tipo",
    ]

    cabecalho = ler_linha_inteira(workbook, nome_aba, linha_cabecalho)
    mapa_colunas: dict[str, int] = {}
    for idx, nome in enumerate(cabecalho, start=1):
        if nome is None:
            continue
        nome_norm = _normalizar_texto(nome)
        for coluna in colunas_identificadoras:
            if nome_norm == _normalizar_texto(coluna):
                mapa_colunas[coluna] = idx

    if not mapa_colunas:
        raise ExcelReaderError(
            "Nenhuma coluna identificadora foi encontrada no cabecalho."
        )

    identificador_norm = _normalizar_texto(identificador)
    resultados: list[dict[str, Any]] = []

    for row_idx, valores in enumerate(
        aba.iter_rows(min_row=linha_cabecalho + 1, values_only=True),
        start=linha_cabecalho + 1,
    ):
        match = False
        for coluna, col_idx in mapa_colunas.items():
            valor = valores[col_idx - 1] if col_idx - 1 < len(valores) else None
            valor_norm = _normalizar_texto(valor)

            if correspondencia_parcial:
                if identificador_norm and identificador_norm in valor_norm:
                    match = True
            else:
                if identificador_norm == valor_norm:
                    match = True

            if match:
                break

        if match:
            dados_linha: dict[str, Any] = {}
            for idx, nome_coluna in enumerate(cabecalho):
                if nome_coluna is None:
                    continue
                chave = str(nome_coluna).strip()
                dados_linha[chave] = valores[idx] if idx < len(valores) else None

            resultados.append({"linha": row_idx, "dados": dados_linha})

    return resultados
