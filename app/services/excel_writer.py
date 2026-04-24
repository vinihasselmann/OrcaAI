"""Escrita segura de dados na planilha real de orcamento (modo append)."""

from __future__ import annotations

import logging
import re
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from shutil import copy2
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

from app.models.schema import MapeamentoExcel
from app.services.field_mapper import ABA_PRINCIPAL, COLUNAS_ZERO_FIXO, MAPPING_RULES

logger = logging.getLogger(__name__)

MAX_EXCEL_ROW = 1_048_576
MIN_EXCEL_ROW_DADOS = 2
CELULA_REGEX = re.compile(r"^([A-Z]+)([1-9]\d*)$")
COLUNAS_CHAVE_APPEND = ("J", "K", "L")


class ExcelWriterError(Exception):
    """Erro de escrita segura em arquivo Excel."""


@dataclass(slots=True)
class ResultadoEscrita:
    """Resumo da operacao de escrita."""

    arquivo_saida: Path
    arquivo_backup: Path | None
    total_mapeamentos: int
    total_escritos: int
    total_ignorados: int
    total_erros: int
    linhas_inseridas: int
    celulas_escritas: int
    erros_encontrados: int


def _obter_colunas_permitidas() -> set[str]:
    """Retorna conjunto de colunas permitidas para escrita na aba principal."""
    base = {"J", "K", "L", "M", "N", "P", "Q", "S", "V", "W", "X", "Y"}
    complementos: set[str] = set()
    complementos.update(COLUNAS_ZERO_FIXO)
    for regra in MAPPING_RULES.values():
        for coluna in regra.get("mapeamento_colunas", {}).values():
            if coluna:
                complementos.add(str(coluna).strip().upper())
    return base | complementos


def criar_backup_planilha(
    caminho_planilha: str | Path,
    pasta_backup: str | Path = "data/output/backups",
) -> Path:
    """Cria backup da planilha original."""
    origem = Path(caminho_planilha)
    if not origem.exists():
        raise ExcelWriterError(f"Planilha nao encontrada para backup: {origem}")
    if origem.suffix.lower() != ".xlsx":
        raise ExcelWriterError(f"Formato invalido para backup: {origem}")

    destino_dir = Path(pasta_backup)
    destino_dir.mkdir(parents=True, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_path = destino_dir / f"{origem.stem}_backup_{timestamp}{origem.suffix}"
    copy2(origem, backup_path)
    logger.info("Backup criado: %s", backup_path)
    return backup_path


def copiar_planilha_base_para_output(
    caminho_planilha_base: str | Path,
    output_dir: str | Path = "data/output",
    nome_arquivo_saida: str | None = None,
) -> Path:
    """Copia a planilha base para um novo arquivo em data/output."""
    origem = Path(caminho_planilha_base)
    if not origem.exists():
        raise ExcelWriterError(f"Planilha base nao encontrada: {origem}")
    if origem.suffix.lower() != ".xlsx":
        raise ExcelWriterError(f"Formato invalido (esperado .xlsx): {origem}")

    destino_dir = Path(output_dir)
    destino_dir.mkdir(parents=True, exist_ok=True)

    if nome_arquivo_saida:
        if not nome_arquivo_saida.lower().endswith(".xlsx"):
            raise ExcelWriterError("nome_arquivo_saida deve terminar com .xlsx")
        nome_saida = nome_arquivo_saida
    else:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        nome_saida = f"{origem.stem}_atualizado_{timestamp}.xlsx"

    destino = destino_dir / nome_saida
    if destino.resolve() == origem.resolve():
        raise ExcelWriterError("Arquivo de saida nao pode sobrescrever o template original.")

    copy2(origem, destino)
    logger.info("Copia de trabalho criada: %s", destino)
    return destino


def localizar_proxima_linha_disponivel(
    caminho_planilha: str | Path,
    nome_aba: str = ABA_PRINCIPAL,
) -> int:
    """Localiza a proxima linha disponivel na aba informada (append)."""
    wb = load_workbook(filename=Path(caminho_planilha), data_only=False)
    try:
        if nome_aba not in wb.sheetnames:
            raise ExcelWriterError(f"Aba nao encontrada: {nome_aba}")
        ws = wb[nome_aba]
        ultima = MIN_EXCEL_ROW_DADOS - 1
        for row_num in range(MIN_EXCEL_ROW_DADOS, ws.max_row + 1):
            if any(
                ws[f"{col}{row_num}"].value is not None
                and str(ws[f"{col}{row_num}"].value).strip()
                for col in COLUNAS_CHAVE_APPEND
            ):
                ultima = row_num
        return ultima + 1
    finally:
        wb.close()


def _extrair_coluna(mapeamento: MapeamentoExcel) -> str:
    """Extrai coluna destino de um mapeamento."""
    if mapeamento.coluna_destino:
        return str(mapeamento.coluna_destino).strip().upper()
    if mapeamento.celula_destino:
        match = CELULA_REGEX.match(str(mapeamento.celula_destino).strip().upper())
        if match:
            return match.group(1)
    raise ExcelWriterError(
        f"Mapeamento sem coluna valida: campo={mapeamento.campo_origem}, "
        f"destino={mapeamento.celula_destino or mapeamento.coluna_destino}"
    )


def _celula_tem_formula(valor: Any) -> bool:
    """Verifica se valor da celula representa formula Excel."""
    return isinstance(valor, str) and valor.startswith("=")


def escrever_dados_mapeados(
    caminho_planilha_copia: str | Path,
    mapeamentos: list[MapeamentoExcel],
    *,
    nome_aba: str = ABA_PRINCIPAL,
) -> tuple[int, int, int, int]:
    """Escreve dados mapeados na planilha copiada em modo append.

    Retorna:
    - linhas_inseridas
    - celulas_escritas
    - ignorados
    - erros
    """
    if not isinstance(mapeamentos, list):
        raise ExcelWriterError("Parametro 'mapeamentos' deve ser uma lista.")

    caminho = Path(caminho_planilha_copia)
    wb = load_workbook(filename=caminho, data_only=False)
    try:
        if nome_aba not in wb.sheetnames:
            raise ExcelWriterError(f"Aba obrigatoria nao encontrada: {nome_aba}")
        ws = wb[nome_aba]

        colunas_permitidas = _obter_colunas_permitidas()
        proxima_linha = localizar_proxima_linha_disponivel(caminho, nome_aba=nome_aba)

        # Mapeia "linha origem do mapeamento" -> "linha append real".
        linha_origem_para_destino: dict[Any, int] = {}
        linhas_com_escrita: set[int] = set()

        celulas_escritas = 0
        ignorados = 0
        erros = 0

        for idx, item in enumerate(mapeamentos, start=1):
            try:
                if not item.permitido_escrever:
                    ignorados += 1
                    continue

                if item.aba_destino != nome_aba:
                    # Protecao: nao escrever em abas de resumo/consolidacao
                    ignorados += 1
                    logger.warning(
                        "Mapeamento ignorado por aba nao permitida: idx=%s aba=%s",
                        idx,
                        item.aba_destino,
                    )
                    continue

                coluna = _extrair_coluna(item)
                if coluna not in colunas_permitidas:
                    ignorados += 1
                    logger.warning(
                        "Coluna nao permitida ignorada: idx=%s coluna=%s campo=%s",
                        idx,
                        coluna,
                        item.campo_origem,
                    )
                    continue

                chave_linha_origem = item.linha_destino if item.linha_destino is not None else f"auto-{idx}"
                if chave_linha_origem not in linha_origem_para_destino:
                    linha_origem_para_destino[chave_linha_origem] = proxima_linha
                    proxima_linha += 1

                linha_destino = linha_origem_para_destino[chave_linha_origem]
                if linha_destino < MIN_EXCEL_ROW_DADOS or linha_destino > MAX_EXCEL_ROW:
                    raise ExcelWriterError(f"Linha fora da area valida: {linha_destino}")

                celula_ref = f"{coluna}{linha_destino}"
                celula = ws[celula_ref]
                anterior = celula.value
                if _celula_tem_formula(anterior):
                    ignorados += 1
                    logger.warning(
                        "Celula com formula preservada: aba=%s celula=%s registro=%s",
                        nome_aba,
                        celula_ref,
                        item.identificador_registro,
                    )
                    continue

                celula.value = item.valor_convertido
                celulas_escritas += 1
                linhas_com_escrita.add(linha_destino)
                logger.info(
                    "Escrita: aba=%s celula=%s anterior=%r novo=%r origem=%s",
                    nome_aba,
                    celula_ref,
                    anterior,
                    item.valor_convertido,
                    item.identificador_registro,
                )
            except Exception as exc:
                erros += 1
                logger.error(
                    "Erro ao escrever mapeamento idx=%s campo=%s registro=%s erro=%s",
                    idx,
                    getattr(item, "campo_origem", None),
                    getattr(item, "identificador_registro", None),
                    exc,
                )

        wb.save(caminho)
        return len(linhas_com_escrita), celulas_escritas, ignorados, erros
    finally:
        wb.close()


def aplicar_mapeamentos_excel(
    caminho_planilha: str | Path,
    mapeamentos: list[MapeamentoExcel],
    *,
    output_dir: str | Path = "data/output",
    criar_backup: bool = True,
    nome_arquivo_saida: str | None = None,
) -> ResultadoEscrita:
    """Fluxo principal: backup opcional, copia template, append seguro e resumo."""
    origem = Path(caminho_planilha)
    if not origem.exists():
        raise ExcelWriterError(f"Planilha nao encontrada: {origem}")
    if origem.suffix.lower() != ".xlsx":
        raise ExcelWriterError(f"Formato invalido (esperado .xlsx): {origem}")

    arquivo_backup: Path | None = None
    if criar_backup:
        arquivo_backup = criar_backup_planilha(origem)

    arquivo_saida = copiar_planilha_base_para_output(
        caminho_planilha_base=origem,
        output_dir=output_dir,
        nome_arquivo_saida=nome_arquivo_saida,
    )

    linhas_inseridas, celulas_escritas, ignorados, erros = escrever_dados_mapeados(
        caminho_planilha_copia=arquivo_saida,
        mapeamentos=mapeamentos,
        nome_aba=ABA_PRINCIPAL,
    )

    logger.info(
        "Resumo escrita: saida=%s linhas_inseridas=%s celulas_escritas=%s ignorados=%s erros=%s",
        arquivo_saida,
        linhas_inseridas,
        celulas_escritas,
        ignorados,
        erros,
    )

    return ResultadoEscrita(
        arquivo_saida=arquivo_saida,
        arquivo_backup=arquivo_backup,
        total_mapeamentos=len(mapeamentos),
        total_escritos=celulas_escritas,
        total_ignorados=ignorados,
        total_erros=erros,
        linhas_inseridas=linhas_inseridas,
        celulas_escritas=celulas_escritas,
        erros_encontrados=erros,
    )
