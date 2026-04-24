from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from app.models.schema import DocumentoImportado, MapeamentoExcel, PecaGenerica
from app.services.excel_writer import aplicar_mapeamentos_excel
from app.services.field_mapper import mapear_documento_para_excel
from tests.conftest import criar_template_base


def test_escrita_segura_excel_append_sem_sobrescrever_formula(tmp_path: Path) -> None:
    template = criar_template_base(tmp_path / "orcamento_base.xlsx")
    output_dir = tmp_path / "output"

    # Coloca formula na proxima linha para validar bloqueio.
    wb = load_workbook(template)
    ws = wb["2A. Lista de Peças"]
    ws["K3"] = "=A1+B1"
    wb.save(template)
    wb.close()

    mapeamentos = [
        MapeamentoExcel(
            aba_destino="2A. Lista de Peças",
            linha_destino=999,  # writer usa append seguro e ignora sobrescrita direta
            coluna_destino="J",
            celula_destino="J999",
            campo_origem="codigo_montagem",
            valor_convertido="CM-NOVO",
            identificador_registro="CM-NOVO",
            permitido_escrever=True,
        ),
        MapeamentoExcel(
            aba_destino="2A. Lista de Peças",
            linha_destino=999,
            coluna_destino="K",
            celula_destino="K999",
            campo_origem="modelo",
            valor_convertido="MOD-NOVO",
            identificador_registro="CM-NOVO",
            permitido_escrever=True,
        ),
        MapeamentoExcel(
            aba_destino="2A. Lista de Peças",
            linha_destino=999,
            coluna_destino="D",  # nao permitida
            celula_destino="D999",
            campo_origem="campo_invalido",
            valor_convertido="NAO",
            identificador_registro="CM-NOVO",
            permitido_escrever=True,
        ),
    ]

    resultado = aplicar_mapeamentos_excel(
        caminho_planilha=template,
        mapeamentos=mapeamentos,
        output_dir=output_dir,
        criar_backup=False,
    )

    assert resultado.celulas_escritas == 2
    assert resultado.linhas_inseridas == 1
    assert resultado.total_ignorados >= 1
    assert resultado.arquivo_saida.exists()

    wb_out = load_workbook(resultado.arquivo_saida, data_only=False)
    ws_out = wb_out["2A. Lista de Peças"]
    # Como K3 possui formula, a proxima linha disponivel para append e a 4.
    assert ws_out["J4"].value == "CM-NOVO"
    assert ws_out["K4"].value == "MOD-NOVO"
    assert ws_out["K3"].value == "=A1+B1"  # formula preservada
    wb_out.close()


def test_escrita_preenche_colunas_zero_fixo_e_laje_vao_em_s(tmp_path: Path) -> None:
    template = criar_template_base(tmp_path / "orcamento_base.xlsx")
    output_dir = tmp_path / "output"

    registro = PecaGenerica(
        arquivo_origem="arquivo.txt",
        tipo_arquivo="geral_pecas_genericas",
        codigo_montagem="CM-G-001",
        modelo="MOD-G",
        marca_tipo="TIPO-G",
        laje_vao_m=6,
        taxa_ca_kg_m3=110,
        taxa_cp_kg_m3=75,
        quantidade=3,
        comprimento_total_m=18,
        area_total_m2=6,
        volume_total_m3=2.2,
    )
    documento = DocumentoImportado(
        arquivo_origem="arquivo.txt",
        tipo_arquivo="geral_pecas_genericas",
        cabecalhos_originais=["codigo_montagem"],
        registros=[registro],
    )
    mapeamentos = mapear_documento_para_excel(
        documento, {"proxima_linha_por_aba": {"2A. Lista de PeÃ§as": 3}}
    )

    resultado = aplicar_mapeamentos_excel(
        caminho_planilha=template,
        mapeamentos=mapeamentos,
        output_dir=output_dir,
        criar_backup=False,
    )

    wb_out = load_workbook(resultado.arquivo_saida, data_only=False)
    ws_out = wb_out["2A. Lista de PeÃ§as"]
    assert ws_out["S3"].value == 6
    assert ws_out["Z3"].value == 0
    assert ws_out["AA3"].value == 0
    assert ws_out["AB3"].value == 0
    assert ws_out["AC3"].value == 0
    assert ws_out["AD3"].value == 0
    assert ws_out["AI3"].value == 0
    assert ws_out["AP3"].value == 0
    wb_out.close()
