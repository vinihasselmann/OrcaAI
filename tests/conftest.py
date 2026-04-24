from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook


def criar_template_base(caminho: Path) -> Path:
    """Cria template minimo com aba '2A. Lista de Peças'."""
    wb = Workbook()
    ws = wb.active
    ws.title = "2A. Lista de Peças"

    # Cabecalhos-base usados pelo sistema.
    ws["B1"] = "quadrante_montagem"
    ws["J1"] = "codigo_montagem"
    ws["K1"] = "modelo"
    ws["L1"] = "marca_tipo"
    ws["M1"] = "largura_preo_m"
    ws["N1"] = "altura_preo_m"
    ws["P1"] = "taxa_ca_kg_m3"
    ws["Q1"] = "taxa_cp_kg_m3"
    ws["S1"] = "comprimento_maximo_m"
    ws["V1"] = "quantidade"
    ws["W1"] = "comprimento_total_m"
    ws["X1"] = "area_total_m2"
    ws["Y1"] = "volume_total_m3"

    # Uma linha inicial para forcar append na linha seguinte.
    ws["J2"] = "EXISTENTE-001"
    ws["K2"] = "MOD-BASE"
    ws["L2"] = "TIPO-BASE"
    ws["V2"] = 1
    ws["W2"] = 1.0
    ws["X2"] = 1.0
    ws["Y2"] = 1.0

    caminho.parent.mkdir(parents=True, exist_ok=True)
    wb.save(caminho)
    wb.close()
    return caminho


def conteudo_txt_geral_pecas() -> str:
    return "\n".join(
        [
            "Relatorio Geral Pecas",
            "codigo_montagem;modelo;marca_tipo;largura_preo_m;altura_preo_m;taxa_ca_kg_m3;taxa_cp_kg_m3;comprimento_maximo_m;quantidade;comprimento_total_m;area_total_m2;volume_total_m3",
            "CM-001;MOD-01;TIPO-A;0,30;0,50;120,0;80,0;12,5;2;25,0;7,5;3,75",
        ]
    )


def conteudo_txt_geral_pecas_auxiliares() -> str:
    return "\n".join(
        [
            "0.Geral Peças Auxiliares;;;;;;;;;;;;",
            "PEÇA-Material;PEÇA-Quadrante de Montagem;Código de montagem;Modelo;Marca de Tipo;PEÇA-Largura Preo (m);PEÇA-Altura Preo (m);TAXA-CA (kg/m³);TAXA-CP (kg/m³);Contagem;PEÇA-Comprimento Total (m);PEÇA-Área Total (m²);PEÇA-Volume Total (m³)",
            ";;;;;;;;;;;;",
            "Pré-moldado;CONCESSIONÁRIA;CP;I;CP01;0,4;0,65;200;0;32;120,8;78,52;1,008",
            "Pré-moldado;CONCESSIONÁRIA;CP;I;CP02;0,4;0,9;400;0;24;241,8;217,62;1,008",
            "Total geral: 28;;;;;;;;;56;362,61;296,15;2,016",
        ]
    )


def conteudo_txt_geral_pecas_genericas() -> str:
    return "\n".join(
        [
            "Relatorio Geral Pecas Genericas",
            "codigo_montagem;modelo;marca_tipo;largura_preo_m;altura_preo_m;espessura_equivalente_cm;distribuicao_cabos;taxa_ca_kg_m3;taxa_cp_kg_m3;laje_vao_m;volume_preenchimento_alveolo_m3;quantidade;comprimento_total_m;area_total_m2;volume_total_m3",
            "CM-G-001;MOD-G;TIPO-G;0,30;0,55;12,5;653  ...  1006,99;110,0;75,0;6,0;0,45;3;18,0;6,0;2,2",
        ]
    )


def conteudo_txt_geral_pecas_alveolares() -> str:
    return "\n".join(
        [
            "Relatorio Geral Pecas Alveolares",
            "codigo_montagem;modelo;marca_tipo;variacao_comprimento_cm;taxa_ca_kg_m3;taxa_cp_kg_m3;comprimento_maximo_m;quantidade;comprimento_total_m;area_total_m2;volume_total_m3",
            "CM-A-001;MOD-A;TIPO-A;10 a 15;100,0;70,0;14,0;4;56,0;14,0;5,0",
        ]
    )
